"""Extract text from incoming documents (PDF / Excel / CSV / Word / plain text)."""
import csv
import io

_MAX_CHARS = 12000


def _pdf(data: bytes) -> str:
    from pypdf import PdfReader

    reader = PdfReader(io.BytesIO(data))
    return "\n".join((p.extract_text() or "") for p in reader.pages)


def _xlsx(data: bytes) -> str:
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    out = []
    for ws in wb.worksheets:
        out.append(f"# גיליון: {ws.title}")
        for row in ws.iter_rows(values_only=True):
            cells = [str(c) for c in row if c is not None]
            if cells:
                out.append(" | ".join(cells))
    return "\n".join(out)


def _docx(data: bytes) -> str:
    from docx import Document

    doc = Document(io.BytesIO(data))
    return "\n".join(p.text for p in doc.paragraphs if p.text.strip())


def _csv(data: bytes) -> str:
    text = data.decode("utf-8", errors="replace")
    rows = list(csv.reader(io.StringIO(text)))
    return "\n".join(" | ".join(r) for r in rows if any(r))


def extract(data: bytes, filename: str, mime: str = "") -> str | None:
    name = (filename or "").lower()
    m = (mime or "").lower()
    try:
        if name.endswith(".pdf") or "pdf" in m:
            text = _pdf(data)
        elif name.endswith((".xlsx", ".xlsm")) or "spreadsheet" in m or "excel" in m:
            text = _xlsx(data)
        elif name.endswith(".docx") or "wordprocessingml" in m:
            text = _docx(data)
        elif name.endswith(".csv") or "csv" in m:
            text = _csv(data)
        elif name.endswith((".txt", ".md")) or m.startswith("text/"):
            text = data.decode("utf-8", errors="replace")
        else:
            return None
    except Exception as e:  # noqa: BLE001
        print(f"[documents] extract failed for {filename}: {e}")
        return None
    text = (text or "").strip()
    if not text:
        return None
    return text[:_MAX_CHARS]
